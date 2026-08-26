"""
ETL Pipeline & Metric Calculator Engine for DOI Monitoring Dashboard (MNJ & KX Principal)
Calculates Stok Max (Qty & Value) dynamically as (DOI Max Master / 30.0) * Avg Sales, and weighted DOI Max per GB.
"""

import os
import openpyxl
import datetime
import pickle
import hashlib
from typing import Dict, List, Any, Optional, Set, Union

def parse_year_month(val: Any) -> Optional[str]:
    """Converts datetime or date string into YYYY-MM string."""
    if not val:
        return None
    if isinstance(val, (datetime.datetime, datetime.date)):
        return val.strftime("%Y-%m")
    val_str = str(val).strip()
    if len(val_str) >= 7 and val_str[0:4].isdigit() and val_str[5:7].isdigit():
        return val_str[0:7]
    return None

def parse_multi_param(val: Any) -> Set[str]:
    """Parses multi-select parameter into a set of normalized strings."""
    if not val:
        return set()
    if isinstance(val, (list, tuple, set)):
        items = {str(x).strip() for x in val if str(x).strip() and str(x).strip().lower() != "all"}
        return items
    val_str = str(val).strip()
    if not val_str or val_str.lower() == "all":
        return set()
    return {x.strip() for x in val_str.split(",") if x.strip() and x.strip().lower() != "all"}

class DataEngine:
    def __init__(self, base_dir: str):
        self.base_dir = os.path.abspath(base_dir)
        self.master_file = os.path.join(self.base_dir, "Master produk.xlsx")
        self.mnj_file = os.path.join(base_dir, "Stok Akhir bulan MNJ.xlsx")
        self.kx_file = os.path.join(base_dir, "Stok Akhir bulan KX dan Produksi.xlsx")
        self.sales_file = os.path.join(base_dir, "Data sales.xlsx")

        self.master_products: Dict[str, Dict[str, Any]] = {}
        self.old_code_map: Dict[str, str] = {}
        self.principal_code_map: Dict[str, str] = {}
        self.available_periods: List[str] = []

        self._is_preloaded = False
        self._mnj_cache: Dict[str, Dict[str, Dict[str, float]]] = {}  # {period: {pcode: {baik, bdp, total}}}
        self._kx_cache: Dict[str, Dict[str, float]] = {}             # {period: {pcode: saldo_akhir_qty}}
        self._sales_cache: Dict[str, Dict[str, float]] = {}          # {month: {pcode: qty}}

    def load_master_data(self) -> Dict[str, Dict[str, Any]]:
        """Reads Master produk.xlsx and builds lookup tables with exact column mapping."""
        if self.master_products:
            return self.master_products

        if not os.path.exists(self.master_file):
            raise FileNotFoundError(f"Master file not found: {self.master_file}")

        wb = openpyxl.load_workbook(self.master_file, read_only=True, data_only=True)
        sheet = wb.active

        # Master Headers:
        # 0: Principal_product_code, 1: Principal_product_code_lama, 2: Product_code,
        # 3: Product_code_lama, 4: Product_name, 5: GB, 6: Harga Dasar,
        # 7: KATEGORI, 8: Keterangan produk, 9: Batch Yield, 10: Line Produksi,
        # 11: Min DOI, 12: Max DOI
        for row in sheet.iter_rows(min_row=2, values_only=True):
            if not row or len(row) < 7 or not row[2]:
                continue

            principal_code = str(row[0]).strip() if row[0] is not None else ""
            principal_code_old = str(row[1]).strip() if row[1] is not None else ""
            product_code = str(row[2]).strip()
            old_code = str(row[3]).strip() if row[3] is not None else ""
            product_name = str(row[4]).strip() if row[4] is not None else ""
            gb = str(row[5]).strip() if row[5] is not None else ""
            
            try:
                harga_dasar = float(row[6]) if row[6] is not None else 0.0
            except (ValueError, TypeError):
                harga_dasar = 0.0

            kategori = str(row[7]).strip() if len(row) > 7 and row[7] is not None else ""
            keterangan = str(row[8]).strip() if len(row) > 8 and row[8] is not None else "Regular"
            if not keterangan:
                keterangan = "Regular"

            try:
                doi_min = float(row[11]) if len(row) > 11 and row[11] is not None else 30.0
            except (ValueError, TypeError):
                doi_min = 30.0

            try:
                doi_max = float(row[12]) if len(row) > 12 and row[12] is not None else 60.0
            except (ValueError, TypeError):
                doi_max = 60.0

            product_info = {
                "principal_code": principal_code,
                "principal_code_old": principal_code_old,
                "product_code": product_code,
                "old_code": old_code,
                "product_name": product_name,
                "gb": gb if gb else "Unassigned",
                "harga_dasar": harga_dasar,
                "kategori": kategori,
                "keterangan": keterangan,
                "doi_min_days": doi_min,
                "doi_max_days": doi_max,
                "target_doi_days": doi_max
            }

            self.master_products[product_code] = product_info
            
            if old_code:
                self.old_code_map[old_code] = product_code
            if principal_code:
                self.principal_code_map[principal_code] = product_code
            if principal_code_old:
                self.principal_code_map[principal_code_old] = product_code

        wb.close()
        return self.master_products

    def resolve_product_code(self, raw_code: str) -> Optional[str]:
        """Resolves raw/old/principal code to primary Product_code."""
        code = raw_code.strip()
        if code in self.master_products:
            return code
        if code in self.old_code_map:
            return self.old_code_map[code]
        if code in self.principal_code_map:
            return self.principal_code_map[code]
        return None

    def _get_cache_checksum(self) -> str:
        h = hashlib.md5()
        for fp in [self.master_file, self.mnj_file, self.kx_file, self.sales_file]:
            abs_fp = os.path.abspath(fp)
            if os.path.exists(abs_fp):
                stat = os.stat(abs_fp)
                h.update(f"{abs_fp}:{stat.st_mtime}:{stat.st_size}".encode("utf-8"))
        return h.hexdigest()

    def preload_all_data(self):
        """Preloads MNJ stock, KX stock, and Sales datasets into memory (with disk caching)."""
        if self._is_preloaded:
            return

        cache_file = os.path.join(self.base_dir, "backend", ".data_cache.pkl")
        current_checksum = self._get_cache_checksum()

        if os.path.exists(cache_file):
            try:
                with open(cache_file, "rb") as f:
                    cache_data = pickle.load(f)
                if cache_data.get("checksum") == current_checksum:
                    print(f"[ENGINE] Loading cached dataset from {cache_file}...", flush=True)
                    self.master_products = cache_data["master_products"]
                    self.old_code_map = cache_data["old_code_map"]
                    self.principal_code_map = cache_data["principal_code_map"]
                    self.available_periods = cache_data["available_periods"]
                    self._mnj_cache = cache_data["_mnj_cache"]
                    self._kx_cache = cache_data["_kx_cache"]
                    self._sales_cache = cache_data["_sales_cache"]
                    self._is_preloaded = True
                    print(f"[ENGINE] Fast cache load complete! {len(self.master_products)} products, {len(self.available_periods)} periods.", flush=True)
                    return
            except Exception as e:
                print(f"[ENGINE] Cache load failed ({e}), re-parsing Excel files...", flush=True)

        print("[ENGINE] Preloading Excel datasets into memory...", flush=True)
        self.load_master_data()

        # 1. Preload MNJ Stock
        if os.path.exists(self.mnj_file):
            wb = openpyxl.load_workbook(self.mnj_file, read_only=True, data_only=True)
            sheet = wb.active
            for row in sheet.iter_rows(min_row=2, values_only=True):
                if not row or len(row) < 10 or not row[3]:
                    continue
                period = parse_year_month(row[9])
                if not period:
                    continue
                target_code = self.resolve_product_code(str(row[3]))
                if not target_code:
                    continue

                qty_baik = float(row[6]) if row[6] is not None else 0.0
                qty_bdp = float(row[8]) if row[8] is not None else 0.0

                if period not in self._mnj_cache:
                    self._mnj_cache[period] = {}
                if target_code not in self._mnj_cache[period]:
                    self._mnj_cache[period][target_code] = {"baik": 0.0, "bdp": 0.0, "total": 0.0}

                self._mnj_cache[period][target_code]["baik"] += qty_baik
                self._mnj_cache[period][target_code]["bdp"] += qty_bdp
                self._mnj_cache[period][target_code]["total"] += (qty_baik + qty_bdp)
            wb.close()

        # 2. Preload KX Principal Stock
        if os.path.exists(self.kx_file):
            wb = openpyxl.load_workbook(self.kx_file, read_only=True, data_only=True)
            sheet = wb.active
            for row in sheet.iter_rows(min_row=4, values_only=True):
                if not row or len(row) < 17 or row[2] is None or row[1] is None:
                    continue
                raw_code = str(row[2]).strip()
                period = parse_year_month(row[16])
                if not period:
                    continue
                saldo_akhir = float(row[15]) if row[15] is not None else 0.0

                target_code = self.resolve_product_code(raw_code)
                if not target_code:
                    continue

                if period not in self._kx_cache:
                    self._kx_cache[period] = {}
                self._kx_cache[period][target_code] = self._kx_cache[period].get(target_code, 0.0) + saldo_akhir
            wb.close()

            # Ensure any negative KX stock balances are reset to 0.0
            for p in self._kx_cache:
                for k in self._kx_cache[p]:
                    if self._kx_cache[p][k] < 0.0:
                        self._kx_cache[p][k] = 0.0

        # 3. Preload Sales Data
        if os.path.exists(self.sales_file):
            wb = openpyxl.load_workbook(self.sales_file, read_only=True, data_only=True)
            sheet = wb.active
            for row in sheet.iter_rows(min_row=2, values_only=True):
                if not row or len(row) < 10 or not row[7]:
                    continue
                month = parse_year_month(row[0])
                if not month:
                    continue
                target_code = self.resolve_product_code(str(row[7]))
                if not target_code:
                    continue

                qty = float(row[9]) if row[9] is not None else 0.0
                if month not in self._sales_cache:
                    self._sales_cache[month] = {}
                self._sales_cache[month][target_code] = self._sales_cache[month].get(target_code, 0.0) + qty
            wb.close()

        mnj_periods = set(self._mnj_cache.keys())
        kx_periods = set(self._kx_cache.keys())
        combined_periods = sorted(list(mnj_periods.union(kx_periods)), reverse=True)

        self.available_periods = combined_periods
        self._is_preloaded = True
        print(f"[ENGINE] Preload complete! {len(self.master_products)} products, {len(self.available_periods)} periods.", flush=True)

        # Save to disk cache for instantaneous future server startups
        try:
            cache_data = {
                "checksum": current_checksum,
                "master_products": self.master_products,
                "old_code_map": self.old_code_map,
                "principal_code_map": self.principal_code_map,
                "available_periods": self.available_periods,
                "_mnj_cache": self._mnj_cache,
                "_kx_cache": self._kx_cache,
                "_sales_cache": self._sales_cache
            }
            with open(cache_file, "wb") as f:
                pickle.dump(cache_data, f)
            print(f"[ENGINE] Saved cache to {cache_file}", flush=True)
        except Exception as e:
            print(f"[ENGINE] Failed to save cache: {e}", flush=True)

    def get_available_periods(self) -> List[str]:
        if not self._is_preloaded:
            self.preload_all_data()
        return self.available_periods

    def get_keterangan_options(self) -> List[str]:
        if not self._is_preloaded:
            self.preload_all_data()
        kets = {p["keterangan"] for p in self.master_products.values() if p.get("keterangan")}
        return sorted(list(kets))

    def get_period_months_window(self, target_period: str, avg_months: int) -> List[str]:
        try:
            year, month = int(target_period[0:4]), int(target_period[5:7])
        except (ValueError, IndexError):
            return [target_period]

        window = []
        cur_y, cur_m = year, month
        for _ in range(avg_months):
            window.append(f"{cur_y:04d}-{cur_m:02d}")
            cur_m -= 1
            if cur_m < 1:
                cur_m = 12
                cur_y -= 1
        return window

    def load_sales(self, target_period: Optional[str] = None, avg_months: int = 1) -> Dict[str, float]:
        if not self._is_preloaded:
            self.preload_all_data()
        if not target_period:
            target_period = self.available_periods[0] if self.available_periods else "2026-07"

        window = self.get_period_months_window(target_period, avg_months)
        sales_sum: Dict[str, float] = {}

        for m in window:
            m_sales = self._sales_cache.get(m, {})
            for code, qty in m_sales.items():
                sales_sum[code] = sales_sum.get(code, 0.0) + qty

        months_count = max(1, avg_months)
        avg_sales: Dict[str, float] = {}
        for code, total_qty in sales_sum.items():
            avg_sales[code] = total_qty / months_count

        return avg_sales

    def get_doi_mnj_report(self, period: Optional[str] = None, avg_months: int = 1) -> List[Dict[str, Any]]:
        """Generates comprehensive report containing MNJ stock, KX principal stock, combined stock, DOI Total, and DOI Max Master."""
        if not self._is_preloaded:
            self.preload_all_data()

        periods = self.get_available_periods()
        selected_period = period if period and period in periods else (periods[0] if periods else "2026-07")

        mnj_data = self._mnj_cache.get(selected_period, {})
        kx_data = self._kx_cache.get(selected_period, {})
        avg_sales_dict = self.load_sales(target_period=selected_period, avg_months=avg_months)

        report = []

        for pcode, pinfo in self.master_products.items():
            stok_info = mnj_data.get(pcode, {"baik": 0.0, "bdp": 0.0, "total": 0.0})
            qty_baik = stok_info["baik"]
            qty_bdp = stok_info["bdp"]
            stok_mnj_qty = stok_info["total"]

            stok_kx_qty = max(0.0, kx_data.get(pcode, 0.0))
            stok_total_qty = stok_mnj_qty + stok_kx_qty

            harga_dasar = pinfo["harga_dasar"]
            stok_mnj_value = stok_mnj_qty * harga_dasar
            stok_kx_value = stok_kx_qty * harga_dasar
            stok_total_value = stok_total_qty * harga_dasar

            avg_sales_qty = avg_sales_dict.get(pcode, 0.0)
            avg_sales_value = avg_sales_qty * harga_dasar

            doi_min_days = pinfo["doi_min_days"]
            doi_max_days = pinfo["doi_max_days"]

            # Calculate Stok Min and Stok Max dynamically: (DOI Max / 30.0) * Avg Sales
            stok_min_qty = (doi_min_days / 30.0) * avg_sales_qty
            stok_max_qty = (doi_max_days / 30.0) * avg_sales_qty
            stok_min_value = stok_min_qty * harga_dasar
            stok_max_value = stok_max_qty * harga_dasar

            # DOI MNJ
            if avg_sales_qty > 0:
                doi_mnj = (stok_mnj_qty / avg_sales_qty) * 30.0
            else:
                doi_mnj = 999.0 if stok_mnj_qty > 0 else 0.0

            # DOI KX
            if avg_sales_qty > 0:
                doi_kx = (stok_kx_qty / avg_sales_qty) * 30.0
            else:
                doi_kx = 999.0 if stok_kx_qty > 0 else 0.0

            # DOI Total
            if avg_sales_qty > 0:
                doi_total = (stok_total_qty / avg_sales_qty) * 30.0
            else:
                doi_total = 999.0 if stok_total_qty > 0 else 0.0

            # Health status based on DOI Total vs DOI Min / DOI Max Master
            def get_health_status(doi_val: float) -> str:
                if doi_val < doi_min_days:
                    return "Understock"
                elif doi_val <= doi_max_days:
                    return "Normal"
                else:
                    return "Overstock"

            # Overstock & Understock Variances
            doi_over_days = round(max(0.0, doi_total - doi_max_days), 1) if doi_total > doi_max_days else 0.0
            val_over = round(max(0.0, stok_total_value - stok_max_value), 2) if doi_total > doi_max_days else 0.0
            qty_over = round(max(0.0, stok_total_qty - stok_max_qty), 2) if doi_total > doi_max_days else 0.0

            doi_under_days = round(max(0.0, doi_max_days - doi_total), 1) if doi_total < doi_min_days else 0.0
            val_under = round(max(0.0, stok_max_value - stok_total_value), 2) if doi_total < doi_min_days else 0.0
            qty_under = round(max(0.0, stok_max_qty - stok_total_qty), 2) if doi_total < doi_min_days else 0.0

            if doi_total > doi_max_days:
                selisih_doi = round(doi_total - doi_max_days, 1)
                selisih_val = round(stok_total_value - stok_max_value, 2)
                selisih_qty = round(stok_total_qty - stok_max_qty, 2)
            elif doi_total < doi_min_days:
                selisih_doi = round(doi_total - doi_max_days, 1)
                selisih_val = round(stok_total_value - stok_max_value, 2)
                selisih_qty = round(stok_total_qty - stok_max_qty, 2)
            else:
                selisih_doi = 0.0
                selisih_val = 0.0
                selisih_qty = 0.0

            doi_after_selisih = round(doi_total - selisih_doi, 1)

            report.append({
                "period": selected_period,
                "product_code": pcode,
                "principal_product_code": pinfo["principal_code"],
                "product_name": pinfo["product_name"],
                "gb": pinfo["gb"],
                "category": pinfo["kategori"],
                "keterangan_produk": pinfo["keterangan"],
                "harga_dasar": harga_dasar,

                # Min & Max threshold quantities and values
                "min_qty": round(stok_min_qty, 2),
                "max_qty": round(stok_max_qty, 2),
                "min_value": round(stok_min_value, 2),
                "max_value": round(stok_max_value, 2),

                # DOI Days thresholds
                "doi_min_days": round(doi_min_days, 1),
                "doi_max_days": round(doi_max_days, 1),
                "target_doi_days": round(doi_max_days, 1),

                # MNJ Stock
                "qty_baik": qty_baik,
                "qty_bdp": qty_bdp,
                "stok_mnj_qty": stok_mnj_qty,
                "stok_mnj_value": round(stok_mnj_value, 2),
                "doi_mnj_days": round(doi_mnj, 1),
                "health_status_mnj": get_health_status(doi_mnj),

                # KX Stock
                "stok_kx_qty": stok_kx_qty,
                "stok_kx_value": round(stok_kx_value, 2),
                "doi_kx_days": round(doi_kx, 1),
                "health_status_kx": get_health_status(doi_kx),

                # Total Combined Stock
                "stok_total_qty": stok_total_qty,
                "stok_total_value": round(stok_total_value, 2),
                "doi_total_days": round(doi_total, 1),
                "health_status_total": get_health_status(doi_total),

                # Overstock & Understock Variances
                "doi_overstock_days": doi_over_days,
                "value_overstock": val_over,
                "qty_overstock": qty_over,

                "doi_understock_days": doi_under_days,
                "value_understock": val_under,
                "qty_understock": qty_under,

                "selisih_doi_days": selisih_doi,
                "selisih_value": selisih_val,
                "selisih_qty": selisih_qty,
                "doi_after_selisih": doi_after_selisih,

                # Sales
                "avg_sales_qty": avg_sales_qty,
                "avg_sales_value": round(avg_sales_value, 2)
            })

        return report

    def get_gb_summary_report(self, period: Optional[str] = None, avg_months: int = 6, keterangan: Union[str, List[str]] = "All", unit: str = "value", products: Union[str, List[str]] = "All", health_status: Union[str, List[str]] = "All") -> List[Dict[str, Any]]:
        """Calculates aggregated DOI metrics grouped per GB and Total Consolidated with weighted DOI Max (Days)."""
        report = self.get_doi_mnj_report(period=period, avg_months=avg_months)
        ket_set = parse_multi_param(keterangan)
        prod_set = parse_multi_param(products)
        health_set = parse_multi_param(health_status)

        gb_map: Dict[str, Dict[str, Any]] = {}

        for r in report:
            if ket_set and r["keterangan_produk"] not in ket_set:
                continue
            p_code = r.get("product_code", "")
            p_pcode = r.get("principal_product_code", "")
            p_old = r.get("old_code", "")
            if prod_set and p_code not in prod_set and p_pcode not in prod_set and p_old not in prod_set:
                continue

            if health_set and r["health_status_total"] not in health_set:
                continue

            gb_name = r["gb"] or "Unassigned"
            if gb_name not in gb_map:
                gb_map[gb_name] = {
                    "gb": gb_name,
                    "total_sku": 0,
                    "stok_mnj_qty": 0.0,
                    "stok_mnj_value": 0.0,
                    "stok_kx_qty": 0.0,
                    "stok_kx_value": 0.0,
                    "stok_total_qty": 0.0,
                    "stok_total_value": 0.0,
                    "min_qty_total": 0.0,
                    "max_qty_total": 0.0,
                    "min_value_total": 0.0,
                    "max_value_total": 0.0,
                    "value_overstock_total": 0.0,
                    "qty_overstock_total": 0.0,
                    "value_understock_total": 0.0,
                    "qty_understock_total": 0.0,
                    "selisih_value_sum": 0.0,
                    "selisih_qty_sum": 0.0,
                    "avg_sales_qty": 0.0,
                    "avg_sales_value": 0.0,
                    "understock_count": 0,
                    "normal_count": 0,
                    "overstock_count": 0
                }

            gb_map[gb_name]["total_sku"] += 1
            gb_map[gb_name]["stok_mnj_qty"] += r["stok_mnj_qty"]
            gb_map[gb_name]["stok_mnj_value"] += r["stok_mnj_value"]
            gb_map[gb_name]["stok_kx_qty"] += r["stok_kx_qty"]
            gb_map[gb_name]["stok_kx_value"] += r["stok_kx_value"]
            gb_map[gb_name]["stok_total_qty"] += r["stok_total_qty"]
            gb_map[gb_name]["stok_total_value"] += r["stok_total_value"]
            gb_map[gb_name]["min_qty_total"] += r["min_qty"]
            gb_map[gb_name]["max_qty_total"] += r["max_qty"]
            gb_map[gb_name]["min_value_total"] += r["min_value"]
            gb_map[gb_name]["max_value_total"] += r["max_value"]

            gb_map[gb_name]["value_overstock_total"] += r["value_overstock"]
            gb_map[gb_name]["qty_overstock_total"] += r["qty_overstock"]
            gb_map[gb_name]["value_understock_total"] += r["value_understock"]
            gb_map[gb_name]["qty_understock_total"] += r["qty_understock"]

            gb_map[gb_name]["selisih_value_sum"] += r["selisih_value"]
            gb_map[gb_name]["selisih_qty_sum"] += r["selisih_qty"]

            gb_map[gb_name]["avg_sales_qty"] += r["avg_sales_qty"]
            gb_map[gb_name]["avg_sales_value"] += r["avg_sales_value"]

            status = r["health_status_total"]
            if status == "Understock":
                gb_map[gb_name]["understock_count"] += 1
            elif status == "Normal":
                gb_map[gb_name]["normal_count"] += 1
            elif status == "Overstock":
                gb_map[gb_name]["overstock_count"] += 1

        summary_list = []
        is_value_mode = (unit.lower() == "value")

        for gb_name, d in sorted(gb_map.items()):
            if is_value_mode:
                stok_m = d["stok_mnj_value"]
                stok_k = d["stok_kx_value"]
                stok_t = d["stok_total_value"]
                min_thresh = d["min_value_total"]
                max_thresh = d["max_value_total"]
                selisih_sum = d["selisih_value_sum"]
                sales = d["avg_sales_value"]
            else:
                stok_m = d["stok_mnj_qty"]
                stok_k = d["stok_kx_qty"]
                stok_t = d["stok_total_qty"]
                min_thresh = d["min_qty_total"]
                max_thresh = d["max_qty_total"]
                selisih_sum = d["selisih_qty_sum"]
                sales = d["avg_sales_qty"]

            doi_m = (stok_m / sales * 30.0) if sales > 0 else (999.0 if stok_m > 0 else 0.0)
            doi_k = (stok_k / sales * 30.0) if sales > 0 else (999.0 if stok_k > 0 else 0.0)
            doi_t = (stok_t / sales * 30.0) if sales > 0 else (999.0 if stok_t > 0 else 0.0)
            doi_min_gb = (min_thresh / sales * 30.0) if sales > 0 else 0.0
            doi_max_gb = (max_thresh / sales * 30.0) if sales > 0 else 0.0
            
            # Calculate Selisih DOI = (Jumlah Selisih Stok / Avg Sales) * 30.0
            selisih_doi_gb = (selisih_sum / sales * 30.0) if sales > 0 else 0.0
            doi_after_selisih_gb = doi_t - selisih_doi_gb

            def get_gb_status(stok_val: float) -> str:
                if stok_val < min_thresh:
                    return "Understock"
                elif stok_val <= max_thresh:
                    return "Normal"
                else:
                    return "Overstock"

            status_gb = get_gb_status(stok_t)

            d["doi_mnj_days"] = round(doi_m, 1)
            d["doi_kx_days"] = round(doi_k, 1)
            d["doi_total_days"] = round(doi_t, 1)
            d["doi_min_days"] = round(doi_min_gb, 1)
            d["doi_max_days"] = round(doi_max_gb, 1)
            d["target_doi_days"] = round(doi_max_gb, 1)

            d["selisih_doi_days"] = round(selisih_doi_gb, 1)
            d["selisih_value"] = round(d["selisih_value_sum"], 2)
            d["selisih_qty"] = round(d["selisih_qty_sum"], 2)
            d["doi_after_selisih"] = round(doi_after_selisih_gb, 1)

            d["value_overstock_total"] = round(d["value_overstock_total"], 2)
            d["value_understock_total"] = round(d["value_understock_total"], 2)

            d["health_status_mnj"] = get_gb_status(stok_m)
            d["health_status_kx"] = get_gb_status(stok_k)
            d["health_status_total"] = status_gb

            d["stok_mnj_value"] = round(d["stok_mnj_value"], 2)
            d["stok_kx_value"] = round(d["stok_kx_value"], 2)
            d["stok_total_value"] = round(d["stok_total_value"], 2)
            d["avg_sales_value"] = round(d["avg_sales_value"], 2)
            summary_list.append(d)

        return summary_list

    def get_historical_doi_trend(self, gb: Union[str, List[str]] = "All", keterangan: Union[str, List[str]] = "All", avg_months: int = 6, unit: str = "value", products: Union[str, List[str]] = "All", health_status: str = "All") -> List[Dict[str, Any]]:
        """Calculates DOI trend for MNJ, KX, and Total Combined over all available periods."""
        periods = sorted(self.get_available_periods())
        month_names = ["Jan", "Feb", "Mar", "Apr", "Mei", "Jun", "Jul", "Agu", "Sep", "Okt", "Nov", "Des"]
        
        gb_set = parse_multi_param(gb)
        ket_set = parse_multi_param(keterangan)
        prod_set = parse_multi_param(products)
        is_value_mode = (unit.lower() == "value")

        trend = []
        for p in periods:
            report = self.get_doi_mnj_report(period=p, avg_months=avg_months)
            
            tot_mnj_qty = 0.0
            tot_mnj_val = 0.0
            tot_kx_qty = 0.0
            tot_kx_val = 0.0
            tot_total_qty = 0.0
            tot_total_val = 0.0
            tot_sales_qty = 0.0
            tot_sales_val = 0.0
            sku_cnt = 0

            for r in report:
                if gb_set and r["gb"] not in gb_set:
                    continue
                if ket_set and r["keterangan_produk"] not in ket_set:
                    continue
                p_code = r.get("product_code", "")
                p_pcode = r.get("principal_product_code", "")
                p_old = r.get("old_code", "")
                if prod_set and p_code not in prod_set and p_pcode not in prod_set and p_old not in prod_set:
                    continue
                if health_status != "All" and r.get("health_status_total") != health_status and r.get("health_status_mnj") != health_status:
                    continue

                sku_cnt += 1
                tot_mnj_qty += r["stok_mnj_qty"]
                tot_mnj_val += r["stok_mnj_value"]
                tot_kx_qty += r["stok_kx_qty"]
                tot_kx_val += r["stok_kx_value"]
                tot_total_qty += r["stok_total_qty"]
                tot_total_val += r["stok_total_value"]
                tot_sales_qty += r["avg_sales_qty"]
                tot_sales_val += r["avg_sales_value"]

            if is_value_mode:
                stok_m = tot_mnj_val
                stok_k = tot_kx_val
                stok_t = tot_total_val
                sales = tot_sales_val
            else:
                stok_m = tot_mnj_qty
                stok_k = tot_kx_qty
                stok_t = tot_total_qty
                sales = tot_sales_qty

            doi_m = (stok_m / sales * 30.0) if sales > 0 else (999.0 if stok_m > 0 else 0.0)
            doi_k = (stok_k / sales * 30.0) if sales > 0 else (999.0 if stok_k > 0 else 0.0)
            doi_t = (stok_t / sales * 30.0) if sales > 0 else (999.0 if stok_t > 0 else 0.0)

            parts = p.split("-")
            month_idx = int(parts[1]) - 1
            m_label = month_names[month_idx] if 0 <= month_idx < 12 else parts[1]
            p_label = f"{m_label} {parts[0]}"

            trend.append({
                "period": p,
                "period_label": p_label,
                "total_sku": sku_cnt,

                "stok_mnj_qty": tot_mnj_qty,
                "stok_mnj_value": round(tot_mnj_val, 2),
                "doi_mnj_days": round(doi_m, 1),

                "stok_kx_qty": tot_kx_qty,
                "stok_kx_value": round(tot_kx_val, 2),
                "doi_kx_days": round(doi_k, 1),

                "stok_total_qty": tot_total_qty,
                "stok_total_value": round(tot_total_val, 2),
                "doi_total_days": round(doi_t, 1),

                "avg_sales_qty": tot_sales_qty,
                "avg_sales_value": round(tot_sales_val, 2)
            })

        return trend
